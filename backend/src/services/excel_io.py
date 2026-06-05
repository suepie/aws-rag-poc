"""申請書 Excel の解析と、AI 判定列の追記書き戻し（openpyxl）。

フォーマット規約: doc/samples/README.md / doc/design/confirmation-assistance.md 4。
ヘッダ行を別名で特定し、質問ID/カテゴリ/質問/回答 列を抽出する。
出力は元セルを保持し、右側に AI 列を追記する。
"""
from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

# 列ヘッダの別名（すべて正規化済み: 小文字・空白除去）
QUESTION_ID_ALIASES = {"質問id", "q_id", "qid", "質問番号", "no", "no.", "番号"}
CATEGORY_ALIASES = {"カテゴリ", "category", "分類", "区分"}
QUESTION_ALIASES = {"質問", "question", "項目", "設問", "質問内容", "質問事項"}
ANSWER_ALIASES = {"回答", "answer", "回答内容", "記入欄", "回答欄"}

# 出力で追記する AI 列（順序を保持）
RESULT_HEADERS = ["AI判定", "AI返答案", "AI根拠", "AI参照", "AI信頼度"]
RESULT_WIDTHS = [12, 60, 50, 30, 10]

_HEADER_FILL = PatternFill("solid", fgColor="7030A0")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

MAX_HEADER_SCAN_ROWS = 5


@dataclass
class ColumnMap:
    question: int
    answer: int
    question_id: int | None = None
    category: int | None = None


@dataclass
class QARow:
    row_index: int
    question_id: str
    category: str
    question: str
    answer: str


@dataclass
class ParsedSheet:
    sheet_title: str
    header_row: int
    columns: ColumnMap
    rows: list[QARow]
    last_column: int


def load_wb(data: bytes) -> Workbook:
    return load_workbook(io.BytesIO(data))


def to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def extract_qa(wb: Workbook, layout: dict | None = None) -> ParsedSheet:
    """ワークブックから Q&A 行を抽出する。"""
    sheet_name = (layout or {}).get("sheet")
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    header_row, header_map = _find_header(ws)
    if header_row is None:
        raise ValueError("ヘッダ行（質問列・回答列）が見つかりません。Excel フォーマット規約を確認してください。")

    columns = ColumnMap(
        question=_match(header_map, QUESTION_ALIASES),
        answer=_match(header_map, ANSWER_ALIASES),
        question_id=_match(header_map, QUESTION_ID_ALIASES),
        category=_match(header_map, CATEGORY_ALIASES),
    )

    rows: list[QARow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        question = _cell_str(ws, r, columns.question)
        answer = _cell_str(ws, r, columns.answer)
        if not question and not answer:
            continue  # 空行はスキップ

        qid = _cell_str(ws, r, columns.question_id) or f"R{r}"
        category = _cell_str(ws, r, columns.category)
        rows.append(QARow(row_index=r, question_id=qid, category=category, question=question, answer=answer))

    return ParsedSheet(
        sheet_title=ws.title,
        header_row=header_row,
        columns=columns,
        rows=rows,
        last_column=ws.max_column,
    )


def append_results(wb: Workbook, parsed: ParsedSheet, results: list) -> None:
    """parsed.rows と同順の results を、元シートの右側に追記する（mutate）。

    results[i] は parsed.rows[i] に対応。各要素は以下属性を持つ（reviewer.Verdict）:
      verdict(str) / reply_draft / rationale / reference_titles(list[str]) / confidence
    """
    ws = wb[parsed.sheet_title]
    start = parsed.last_column + 1

    for i, header in enumerate(RESULT_HEADERS):
        col = start + i
        cell = ws.cell(parsed.header_row, col, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = RESULT_WIDTHS[i]

    for row, v in zip(parsed.rows, results, strict=True):
        values = [
            jp_verdict(v.verdict),
            v.reply_draft,
            v.rationale,
            "; ".join(v.reference_titles),
            v.confidence,
        ]
        for i, val in enumerate(values):
            cell = ws.cell(row.row_index, start + i, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


# --- 内部ヘルパー -----------------------------------------------------------

def _find_header(ws) -> tuple[int | None, dict[str, int] | None]:
    """先頭数行を走査し、質問列と回答列を含む行をヘッダとみなす。"""
    scan_to = min(MAX_HEADER_SCAN_ROWS, ws.max_row)
    for r in range(1, scan_to + 1):
        mapping: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            norm = _norm(ws.cell(r, c).value)
            if norm and norm not in mapping:
                mapping[norm] = c
        if _match(mapping, QUESTION_ALIASES) and _match(mapping, ANSWER_ALIASES):
            return r, mapping
    return None, None


def _match(mapping: dict[str, int] | None, aliases: set[str]) -> int | None:
    if not mapping:
        return None
    for alias in aliases:
        if alias in mapping:
            return mapping[alias]
    return None


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("　", "")


def _cell_str(ws, row: int, col: int | None) -> str:
    if not col:
        return ""
    val = ws.cell(row, col).value
    return "" if val is None else str(val).strip()


_JP_VERDICT = {
    "approved": "許可",
    "conditional": "条件付き",
    "needs_review": "要確認",
    "rejected": "却下",
}


def jp_verdict(verdict: str) -> str:
    return _JP_VERDICT.get(verdict, verdict)
