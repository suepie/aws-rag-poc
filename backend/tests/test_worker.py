"""Worker の結合テスト（S3/DynamoDB をモックし、実 Excel I/O で往復）。"""
import io
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))


def _sample_bytes():
    wb = Workbook()
    ws = wb.active
    ws.append(["質問ID", "質問", "回答"])
    for i in range(1, 13):  # 12 問（進捗更新の刻みも検証）
        ws.append([f"Q-{i:03d}", f"質問{i}", f"回答{i}"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def worker(monkeypatch):
    import services.s3_util as s3_util
    import services.job_store as job_store
    import handlers.worker as worker_mod

    store = {"out": None, "completed": None, "progress": []}

    monkeypatch.setattr(s3_util, "get_bytes", lambda key: _sample_bytes())
    monkeypatch.setattr(s3_util, "put_bytes", lambda key, data, ct: store.__setitem__("out", (key, data)))
    monkeypatch.setattr(job_store, "update_progress", lambda jid, current, total: store["progress"].append((current, total)))
    monkeypatch.setattr(job_store, "complete_job", lambda jid, **kw: store.__setitem__("completed", kw))
    monkeypatch.setattr(job_store, "fail_job", lambda jid, err: store.__setitem__("failed", err))
    monkeypatch.setattr(job_store, "get_job", lambda jid: {"job_id": jid, "s3_key_in": "reviews/inbound/x.xlsx"})
    return worker_mod, store


def test_worker_processes_and_appends(worker):
    worker_mod, store = worker
    res = worker_mod.lambda_handler({"job_id": "REV-test"})
    assert res["ok"] is True

    # 出力 Excel に AI 列が追記されている
    assert store["out"] is not None
    key_out, data = store["out"]
    assert key_out.startswith("reviews/outbound/")
    ws = load_workbook(io.BytesIO(data)).active
    assert ws.cell(1, 4).value == "AI判定"
    assert ws.cell(2, 4).value == "要確認"

    # 完了サマリ
    summary = store["completed"]["summary"]
    assert summary["total"] == 12
    assert summary["needs_review"] == 12

    # 進捗は 5 件刻み + 最終
    assert (5, 12) in store["progress"]
    assert (12, 12) in store["progress"]


def test_worker_missing_job_id(worker):
    worker_mod, _ = worker
    assert worker_mod.lambda_handler({})["ok"] is False
