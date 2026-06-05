"""excel_io と reviewer の単体テスト。"""
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from services import excel_io, reviewer  # noqa: E402


def _wb(header, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    return wb


def test_extract_japanese_headers_with_id():
    wb = _wb(
        ["質問ID", "カテゴリ", "質問", "回答"],
        [["Q-001", "暗号化", "暗号化方式は？", "SSE-S3"], ["Q-002", "ログ", "ログは？", "CloudTrail"]],
    )
    parsed = excel_io.extract_qa(wb)
    assert parsed.header_row == 1
    assert len(parsed.rows) == 2
    assert parsed.rows[0].question_id == "Q-001"
    assert parsed.rows[0].category == "暗号化"
    assert parsed.rows[0].question == "暗号化方式は？"
    assert parsed.rows[1].answer == "CloudTrail"


def test_extract_english_aliases_and_fallback_id():
    wb = _wb(
        ["Question", "Answer"],
        [["Where is data stored?", "ap-northeast-1"]],
    )
    parsed = excel_io.extract_qa(wb)
    assert parsed.columns.question_id is None
    assert parsed.rows[0].question_id == "R2"  # フォールバック（行番号）


def test_skip_blank_rows():
    wb = _wb(
        ["質問", "回答"],
        [["q1", "a1"], [None, None], ["", ""], ["q2", "a2"]],
    )
    parsed = excel_io.extract_qa(wb)
    assert [r.question for r in parsed.rows] == ["q1", "q2"]


def test_header_not_found_raises():
    wb = _wb(["foo", "bar"], [["x", "y"]])
    try:
        excel_io.extract_qa(wb)
        raise AssertionError("should have raised")
    except ValueError as exc:
        assert "ヘッダ行" in str(exc)


def test_header_on_second_row():
    wb = Workbook()
    ws = wb.active
    ws.append(["クラウド利用申請書", None])       # タイトル行
    ws.append(["質問", "回答"])                   # ヘッダ行
    ws.append(["q1", "a1"])
    parsed = excel_io.extract_qa(wb)
    assert parsed.header_row == 2
    assert len(parsed.rows) == 1


def test_append_results_round_trip():
    wb = _wb(["質問ID", "質問", "回答"], [["Q-1", "q1", "a1"], ["Q-2", "q2", "a2"]])
    parsed = excel_io.extract_qa(wb)
    results = [
        reviewer.Verdict("needs_review", "返答案", "根拠", "low", ["ref-a"]),
        reviewer.Verdict("approved", "OK", "問題なし", "high", []),
    ]
    excel_io.append_results(wb, parsed, results)

    # bytes 化して再読込（保存往復を検証）
    reloaded = load_workbook(__import__("io").BytesIO(excel_io.to_bytes(wb)))
    ws = reloaded.active
    # 元の 3 列の右（4列目〜）に AI 列が並ぶ
    assert ws.cell(1, 4).value == "AI判定"
    assert ws.cell(1, 5).value == "AI返答案"
    assert ws.cell(1, 8).value == "AI信頼度"
    assert ws.cell(2, 4).value == "要確認"        # needs_review の日本語ラベル
    assert ws.cell(2, 8).value == "low"
    # 元セルは保持
    assert ws.cell(2, 3).value == "a1"


def test_summarize():
    results = [
        reviewer.Verdict("approved", "", ""),
        reviewer.Verdict("needs_review", "", ""),
        reviewer.Verdict("needs_review", "", ""),
    ]
    s = reviewer.summarize(results)
    assert s["approved"] == 1
    assert s["needs_review"] == 2
    assert s["total"] == 3


def test_sample_excel_parses():
    """同梱のサンプル Excel が規約どおり解析できる。"""
    sample = Path(__file__).resolve().parents[2] / "doc" / "samples" / "sample-application.xlsx"
    if not sample.exists():
        return  # 生成前ならスキップ
    parsed = excel_io.extract_qa(excel_io.load_wb(sample.read_bytes()))
    assert len(parsed.rows) == 8
    assert parsed.rows[0].question_id == "Q-001"
