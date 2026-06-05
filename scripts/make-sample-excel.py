#!/usr/bin/env python3
"""確認支援のサンプル申請書 Excel を生成する。

Excel フォーマット規約（[doc/design/confirmation-assistance.md] 4.1）に沿った
入力 Excel の見本。ヘッダ行 + 回答済みの行を持つ。

実行: python3 scripts/make-sample-excel.py
出力: doc/samples/sample-application.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "doc" / "samples" / "sample-application.xlsx"

# (質問ID, カテゴリ, 質問文, 回答)
ROWS = [
    ("Q-001", "概要", "利用するクラウドサービス名を記入してください。", "Amazon S3"),
    ("Q-002", "概要", "サービス提供事業者を記入してください。", "Amazon Web Services"),
    ("Q-003", "データ保管", "データを保管する AWS リージョンを記入してください。",
     "本番は ap-northeast-1。バックアップは us-east-1 にクロスリージョン複製。"),
    ("Q-004", "暗号化", "保管データの暗号化方式を記入してください。",
     "SSE-S3 (AES-256) を有効化。バケットポリシーで非暗号化 PUT を拒否。"),
    ("Q-005", "アクセス制御", "アクセス制御の方式を記入してください。",
     "IAM ロールで最小権限を付与。MFA を必須化。"),
    ("Q-006", "ログ", "操作ログの取得方法を記入してください。",
     "CloudTrail と S3 アクセスログを有効化し、90 日間保管。"),
    ("Q-007", "バックアップ", "バックアップ方式を記入してください。",
     "日次でバージョニング + クロスリージョン複製。"),
    ("Q-008", "委託", "再委託先の有無と管理方法を記入してください。",
     "再委託なし。"),
]

HEADER = ["質問ID", "カテゴリ", "質問", "回答"]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "申請書"

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(vertical="center")

    for r, row in enumerate(ROWS, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [10, 12, 46, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
